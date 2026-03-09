import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import zipfile

st.set_page_config(page_title="Gerador de Avaliações Práticas", layout="wide")
st.title("Sincronização de Avaliações (Importação -> Modelo)")

# --- CONFIGURAÇÃO DO MODELO (Ajuste aqui) ---
# Onde o nome deve ser escrito no seu ficheiro .xlsm final
CELULA_DESTINO_NOME = 'B5' 
# Onde começam as notas dos parâmetros no seu modelo
LINHA_INICIO_NOTAS_MODELO = 15
COLUNA_NOTAS_MODELO = 'C'

# 1. Upload de Ficheiros
col1, col2 = st.columns(2)
with col1:
    file_import = st.file_uploader("1. Ficheiro 'Importação' (.xlsx)", type=["xlsx"])
with col2:
    file_modelo = st.file_uploader("2. Ficheiro 'Modelo' (.xlsm)", type=["xlsm"])

if file_import and file_modelo:
    try:
        # --- PASSO 1: LER DADOS DE IMPORTAÇÃO ---
        # Lemos a partir da linha 13 (index 12)
        df_import = pd.read_excel(file_import, skiprows=12)
        
        if 'nomefrmo' not in df_import.columns:
            st.error("Erro: Não foi encontrada a coluna 'nomefrmo' no ficheiro de importação.")
        else:
            # Limpar linhas sem nome
            df_import = df_import.dropna(subset=['nomefrmo'])
            st.success(f"Detetados {len(df_import)} formandos.")

            # --- PASSO 2: IDENTIFICAR PARÂMETROS ---
            # Vamos assumir que as colunas após 'nomefrmo' são os parâmetros
            idx_nome = df_import.columns.get_loc('nomefrmo')
            colunas_parametros = df_import.columns[idx_nome + 1:] # Colunas à direita de nomefrmo
            
            st.write("**Colunas de avaliação detetadas:**", list(colunas_parametros))

            if st.button("Gerar Ficheiros XLSM"):
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                    bar = st.progress(0)
                    
                    for idx, linha in df_import.iterrows():
                        # Preparar o Modelo (Com Macros)
                        file_modelo.seek(0)
                        wb = load_workbook(file_modelo, keep_vba=True)
                        ws = wb.active
                        
                        # Inserir Nome do Formando
                        ws[CELULA_DESTINO_NOME] = linha['nomefrmo']
                        
                        # Inserir Notas dos Parâmetros
                        for i, col_nome in enumerate(colunas_parametros):
                            valor_nota = linha[col_nome]
                            if pd.notna(valor_nota):
                                # Escreve na coluna C, linhas 15, 16, 17...
                                celula_alvo = f"{COLUNA_NOTAS_MODELO}{LINHA_INICIO_NOTAS_MODELO + i}"
                                ws[celula_alvo] = valor_nota
                        
                        # Guardar em memória
                        output = io.BytesIO()
                        wb.save(output)
                        
                        # Nome do ficheiro individual
                        fn = f"Avaliacao_{str(linha['nomefrmo']).replace(' ', '_')}.xlsm"
                        zip_file.writestr(fn, output.getvalue())
                        
                        bar.progress((idx + 1) / len(df_import))
                
                st.success("✅ Processamento completo!")
                st.download_button(
                    label="📥 Baixar Pasta de Avaliações (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name="Avaliacoes_Praticas.zip",
                    mime="application/zip"
                )

    except Exception as e:
        st.error(f"Ocorreu um erro técnico: {e}")
